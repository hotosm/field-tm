# Copyright (c) Humanitarian OpenStreetMap Team
#
# This file is part of osm_fieldwork.
#
#     osm-fieldwork is free software: you can redistribute it and/or modify
#     it under the terms of the GNU General Public License as published by
#     the Free Software Foundation, either version 3 of the License, or
#     (at your option) any later version.
#
#     osm-fieldwork is distributed in the hope that it will be useful,
#     but WITHOUT ANY WARRANTY; without even the implied warranty of
#     MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#     GNU General Public License for more details.
#
#     You should have received a copy of the GNU General Public License
#     along with osm_fieldwork.  If not, see <https:#www.gnu.org/licenses/>.
#
"""Test functionality of update_form.py."""
import io
from io import BytesIO
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook, worksheet
from pyxform.xls2xform import convert as xform_convert

from osm_fieldwork.update_xlsform import append_field_mapping_fields
from osm_fieldwork.xlsforms import buildings, healthcare
from osm_fieldwork.form_components.translations import INCLUDED_LANGUAGES
from osm_fieldwork.conversion_to_xlsform import convert_to_xlsform

async def test_merge_mandatory_fields():
    """Merge the mandatory fields XLSForm to a test survey form."""
    test_form = Path(__file__).parent / "test_data" / "test_form_for_mandatory_fields.xls"

    with open(test_form, "rb") as xlsform:
        form_bytes = BytesIO(xlsform.read())

    xformid, updated_form = await append_field_mapping_fields(form_bytes, "buildings")
    workbook = load_workbook(filename=BytesIO(updated_form.getvalue()))
    # Write merged xlsform to file for debugging
    with open("merged_xlsform.xlsx", "wb") as merged_xlsform:
        merged_xlsform.write(updated_form.getvalue())

    # remove duplicate field names in 'survey'
    survey = workbook["survey"]
    seen = set()
    for row in range(2, survey.max_row + 1):
        name = survey.cell(row=row, column=2).value
        if name in seen:
            survey.delete_rows(row, 1)
        else:
            seen.add(name)

    check_survey_sheet(workbook)
    check_entities_sheet(workbook)
    check_form_title(workbook)

    # Check it's still a valid xlsform by converting to XML
    xform_convert(updated_form)
    check_translation_fields(workbook)



async def test_add_extra_select_from_file():
    """Append extra select_one_from_file questions based on Entity list names."""
    test_form = Path(__file__).parent / "test_data" / "test_form_for_mandatory_fields.xls"

    with open(test_form, "rb") as xlsform:
        form_bytes = BytesIO(xlsform.read())

    xformid, updated_form = await append_field_mapping_fields(form_bytes, "buildings", additional_entities=["roads", "waterpoints"])
    workbook = load_workbook(filename=BytesIO(updated_form.getvalue()))

    survey_sheet = workbook["survey"]
    name_column = [cell.value for cell in survey_sheet["B"]]

    assert "roads" in name_column, "The 'roads' field was not added to the survey sheet."
    assert "waterpoints" in name_column, "The 'waterpoints' field was not added to the survey sheet."


async def test_buildings_xlsform():
    """Merge and test if buildings form is a valid XLSForm."""
    form_bytes = io.BytesIO(convert_to_xlsform(str(buildings)))
    xformid, updated_form = await append_field_mapping_fields(form_bytes, "buildings")
    # Check it's still a valid xlsform by converting to XML
    xform_convert(updated_form)

    workbook = load_workbook(filename=BytesIO(updated_form.getvalue()))
    check_translation_fields(workbook)


async def test_healthcare_xlsform():
    """Merge and test if buildings form is a valid XLSForm."""
    form_bytes = io.BytesIO(convert_to_xlsform(str(healthcare)))
    xformid, updated_form = await append_field_mapping_fields(form_bytes, "healthcare")
    # Check it's still a valid xlsform by converting to XML
    xform_convert(updated_form)


async def test_odk_collect_entity_task_selection():
    """Test that ODK Collect forms use entity-based task selection."""
    form_bytes = io.BytesIO(convert_to_xlsform(str(buildings)))
    xformid, updated_form = await append_field_mapping_fields(
        form_bytes, "buildings", use_odk_collect=True
    )
    workbook = load_workbook(filename=BytesIO(updated_form.getvalue()))
    survey_sheet = workbook["survey"]

    # Extract column values
    type_col = [cell.value for cell in survey_sheet["A"]]
    name_col = [cell.value for cell in survey_sheet["B"]]

    # Check select_one_from_file tasks.csv is present
    assert "select_one_from_file tasks.csv" in type_col, (
        "select_one_from_file tasks.csv not found in survey types"
    )

    # Check the task field name is 'task' (not 'task_filter')
    task_type_idx = type_col.index("select_one_from_file tasks.csv")
    assert name_col[task_type_idx] == "task", (
        f"Expected field name 'task', got '{name_col[task_type_idx]}'"
    )

    # Check selected_task_id calculated field exists
    assert "selected_task_id" in name_col, (
        "selected_task_id calculated field not found"
    )

    # Check no task_ids in choices sheet
    choices_sheet = workbook["choices"]
    list_name_col = [cell.value for cell in choices_sheet["A"]]
    assert "task_ids" not in list_name_col, (
        "task_ids should not be in choices sheet"
    )

    # Check feature choice_filter references selected_task_id
    feature_idx = name_col.index("feature")
    # Find the choice_filter column
    header = [cell.value for cell in next(survey_sheet.iter_rows(min_row=1, max_row=1))]
    if "choice_filter" in header:
        cf_col_idx = header.index("choice_filter") + 1
        choice_filter_val = survey_sheet.cell(row=feature_idx + 1, column=cf_col_idx).value
        assert "${selected_task_id}" in choice_filter_val, (
            f"Feature choice_filter should reference ${{selected_task_id}}, got: {choice_filter_val}"
        )


async def test_status_values_are_text():
    """Test that status values use text instead of integers."""
    form_bytes = io.BytesIO(convert_to_xlsform(str(buildings)))
    xformid, updated_form = await append_field_mapping_fields(
        form_bytes, "buildings", use_odk_collect=True
    )
    workbook = load_workbook(filename=BytesIO(updated_form.getvalue()))
    survey_sheet = workbook["survey"]

    name_col = [cell.value for cell in survey_sheet["B"]]
    header = [cell.value for cell in next(survey_sheet.iter_rows(min_row=1, max_row=1))]

    # Find calculation column
    calc_col_idx = header.index("calculation") + 1

    # Check status field uses text values
    status_idx = name_col.index("status")
    status_calc = survey_sheet.cell(row=status_idx + 1, column=calc_col_idx).value
    assert "'mapped'" in status_calc, f"Status calculation should use 'mapped', got: {status_calc}"
    assert "'invalid'" in status_calc, f"Status calculation should use 'invalid', got: {status_calc}"

    # Check default is text
    if "default" in header:
        default_col_idx = header.index("default") + 1
        status_default = survey_sheet.cell(row=status_idx + 1, column=default_col_idx).value
        assert status_default == "mapped", f"Status default should be 'mapped', got: {status_default}"

    # Check color calculations use text status
    fill_idx = name_col.index("fill")
    fill_calc = survey_sheet.cell(row=fill_idx + 1, column=calc_col_idx).value
    assert "'unmapped'" in fill_calc, f"Fill calculation should reference 'unmapped', got: {fill_calc}"
    assert "'invalid'" in fill_calc, f"Fill calculation should reference 'invalid', got: {fill_calc}"


def check_survey_sheet(workbook: Workbook) -> None:
    """Check the 'survey' sheet values and ensure no duplicates in 'name' column."""
    survey_sheet = get_sheet(workbook, "survey")
    name_col_index = get_column_index(survey_sheet, "name")
    calculation_col_index = get_column_index(survey_sheet, "calculation")
    check_for_duplicates(survey_sheet, name_col_index)


def check_entities_sheet(workbook: Workbook) -> None:
    """Check the 'entities' sheet values."""
    entities_sheet = get_sheet(workbook, "entities")
    label_col_index = get_column_index(entities_sheet, "label")

    test_label_present = any(
        row[0].value == "test label"
        for row in entities_sheet.iter_rows(min_col=label_col_index, max_col=label_col_index, min_row=2)
    )
    assert not test_label_present, "'test label' found in the 'label' column of 'entities' sheet."


def check_form_title(workbook: Workbook) -> None:
    """Check if the form_title is set correctly in the 'settings' sheet."""
    settings_sheet = get_sheet(workbook, "settings")
    form_title_col_index = get_column_index(settings_sheet, "form_title")

    form_title_value = settings_sheet.cell(row=2, column=form_title_col_index).value
    # NOTE previously we would strip 's' from plurals, but we no longer do this
    assert form_title_value == "buildings", "form_title field is not set to 'building'"


def check_translation_fields(workbook: Workbook):
    """Check if translation fields for all included languages were correctly matched."""
    survey_sheet = workbook["survey"]
    header = [cell.value for cell in next(survey_sheet.iter_rows(min_row=1, max_row=1))]

    if "label" in header: # Allow bare 'label' column without translations
        return

    found_langs = set()
    for col in header:
        if not col or not col.startswith("label::"):
            continue

        match = re.match(r"label::([^(]+)(?:\(([^)]+)\))?", col)
        if not match:
            continue

        lang_key = match.group(1).strip().lower()
        lang_code = match.group(2) or INCLUDED_LANGUAGES.get(lang_key)
        if lang_key in INCLUDED_LANGUAGES and lang_code == INCLUDED_LANGUAGES[lang_key]:
            found_langs.add(lang_key)

    assert found_langs.issubset(INCLUDED_LANGUAGES.keys()), (
        f"Unexpected translation columns: {found_langs - INCLUDED_LANGUAGES.keys()}"
    )


def get_sheet(workbook: Workbook, sheet_name: str) -> worksheet.worksheet.Worksheet:
    """Helper function to get a sheet or raise an error."""
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"The '{sheet_name}' sheet was not found in the workbook")
    return workbook[sheet_name]


def check_for_duplicates(sheet: worksheet.worksheet.Worksheet, col_index: int) -> None:
    """Check for any duplicate values in a specific column of a sheet, ignoring None values."""
    seen_values = set()
    for row in sheet.iter_rows(min_col=col_index, max_col=col_index, min_row=2):
        value = row[0].value
        if value is None:
            # Skip None values, allowing them to appear multiple times
            continue
        if value in seen_values:
            raise AssertionError(f"Duplicate value '{value}' found in column '{col_index}' of sheet '{sheet.title}'.")
        seen_values.add(value)


def get_column_index(sheet: worksheet.worksheet.Worksheet, column_name: str) -> int:
    """Get the column index for the given column name."""
    for col_idx, col in enumerate(sheet.iter_cols(1, sheet.max_column), start=1):
        if col[0].value == column_name:
            return col_idx
    raise ValueError(f"Column '{column_name}' not found.")


def get_row_index(sheet: worksheet.worksheet.Worksheet, column_index: int, value: str) -> int:
    """Get the row index where the given column has the specified value."""
    for row_idx, row in enumerate(sheet.iter_rows(min_col=column_index, max_col=column_index), start=1):
        if row[0].value == value:
            return row_idx
    raise ValueError(f"Value '{value}' not found in column {column_index}.")
